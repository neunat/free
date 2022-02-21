#!/usr/bin/env python

from xml.dom import minidom
import os
import openpyxl
import xml.etree.ElementTree as gfg
import glob

b = open('template.txt', 'r')
bb = b.read()
xml = open('new_database.xml', 'w')

path = "Data"
dir_list = os.listdir(path)

root = gfg.Element("pages")#xml things
tree = gfg.ElementTree(root)




for file in glob.glob("uploads/*.xlsx"):
  path = (file)

print(path)#Needs to open the one file in 'uploads' folder


wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
 
sheet_object = wb_obj.active
max_rows = sheet_object.max_row#finding how many rows there are

os.chdir('Data')
i = 2#if i=1, it would use "21/22" (the year) as well
while i <= max_rows:
  cell_obj = sheet_obj.cell(row = i, column = 1)
  if (cell_obj.value is None) or (cell_obj.value == "GYM") or (cell_obj.value == "DANCE"):#filters out empty, "dance", and "gym" results
    a = 1
  else:
    txt = cell_obj.value
    x = "RM " in txt
    if (x == True):#filters out room number results
      a = 1
    else:
      
        m1 = gfg.Element("school")
        root.append (m1)
        
        b1 = gfg.SubElement(m1, "data")
        b1.text = (cell_obj.value)
        
        print(cell_obj.value)


        searchThis = cell_obj.value
        searchThis = searchThis.replace(" ", "")
        
        #xml.write(xml_str) 
        if searchThis in dir_list:
          #skip, file already exists
          a = 1
          
        else: 
          #create the html file 
          f = open((searchThis), "w")
          #Copy template into file here
          f.write(bb)
          
          
          
          f.close()
          

          
         

  i = i+1
  
b.close()
xml.close()


tree = gfg.ElementTree(root)
tree.write("new_database.xml")





 







#https://www.geeksforgeeks.org/create-xml-documents-using-python/